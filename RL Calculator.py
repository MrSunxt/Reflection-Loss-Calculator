import sys
import os
import time
import cmath
import math
from openpyxl import load_workbook, Workbook
from PyQt5.QtCore import QThread, pyqtSignal, Qt, QVariant
from PyQt5.QtGui import QIcon, QColor, QFont
from PyQt5.QtWidgets import *

path = ''                               # 全局路径
start_thick = 0                         # 起始厚度
end_thick = 10                          # 终止厚度
thick = 10                              # 实际厚度范围
step = 0.1                              # 间隔厚度
thick_count = 0                         # 总点数


class RlCalculate(QThread):
    display_sig = pyqtSignal(str)
    process_sig = pyqtSignal(str)
    path_sig = pyqtSignal(str)
    flag_sig = pyqtSignal(bool)
    except_sig = pyqtSignal(str)
    except_sig2 = pyqtSignal(int)

    def __init__(self):
        super(RlCalculate, self).__init__()

    @staticmethod
    def check_time():
        t = time.perf_counter()
        return t

    # 数据分类
    def sort_data(self):
        e = []
        u = []
        if path.split('.')[-1] == 'xlsx':
            self.display_sig.emit('& 正在读取Excel中的数据......')
            data_wb = load_workbook(path)
            data_ws = data_wb.active
            f = [i.value * 1000000000 for i in data_ws['A']]
            r_e = [i.value for i in data_ws['B']]
            i_e = [i.value for i in data_ws['C']]
            r_u = [i.value for i in data_ws['D']]
            i_u = [i.value for i in data_ws['E']]
            data_wb.close()
            for i in range(0, len(f)):
                e.append(complex(r_e[i], -i_e[i]))
                u.append(complex(r_u[i], -i_u[i]))
            return f, e, u
        else:
            with open(path, 'r') as file:
                data = file.readlines()
            self.display_sig.emit('& 正在读取数据......')
            index = 0
            data = [i for i in data if i != '\n']
            while True:
                if len(data[index]) != len(data[index + 3]):
                    index += 1
                elif len(data[index].split()) != len(data[-1].split()):
                    index += 1
                else:
                    break
            f = []
            r_e = []
            i_e = []
            r_u = []
            i_u = []
            for i in data[index:]:
                if len(data[index].split()[0]) >= 10:
                    f.append(float(i.split()[0]))
                else:
                    f.append(float(i.split()[0]) * 1000000000)
                r_e.append(float(i.split()[1]))
                i_e.append(float(i.split()[2]))
                r_u.append(float(i.split()[3]))
                i_u.append(float(i.split()[4]))
            for i in range(0, len(f)):
                e.append(complex(r_e[i], -i_e[i]))
                u.append(complex(r_u[i], -i_u[i]))
            return f, e, u

    def run(self):
        self.flag_sig.emit(False)
        t1 = self.check_time()
        try:
            data = self.sort_data()
            f = data[0]
            e = data[1]
            u = data[2]
            self.display_sig.emit('& 频率范围：{}-{}GHz'.format(f[0]/1000000000,
                                                           f[-1]/1000000000))
            self.display_sig.emit(
                '& 厚度范围：{}-{}mm，间隔{}mm\n'
                '& 共计 {} 个数据点'.format(start_thick, end_thick, step, thick_count))
            self.display_sig.emit('& 初始化结果文件......')
            item = ['frequency']
            for i in range(thick_count):
                item.append(str(round(start_thick + i * step, 2)) + 'mm')
            wb = load_workbook('RL-IM.xlsx')
            rl_ws = wb['RL']
            im_ws = wb['IM']
            wb.remove(rl_ws)
            wb.remove(im_ws)
            ws_rl = wb.create_sheet('RL')
            ws_rl.append(item)
            ws_im = wb.create_sheet('IM')
            ws_im.append(item)
            self.display_sig.emit('& 计算中，请耐心等待结果......')
            for i in range(0, len(f)):
                list_rl = [f[i] / 1000000000]
                list_im = [f[i] / 1000000000]
                for j in range(0, thick_count):
                    d = round((start_thick / 1000) + (step / 1000) * j, 5)
                    zin = cmath.sqrt(u[i] / e[i]) * cmath.tanh(complex(0, 1) * (
                            (2 * cmath.pi * f[i] * d) / 300000000) * cmath.sqrt(
                        u[i] * e[i]))
                    im = abs(zin)
                    rl = 20 * cmath.log10(abs((zin - 1) / (zin + 1)))
                    list_rl.append(rl.real)
                    list_im.append(im)
                ws_rl.append(list_rl)
                ws_im.append(list_im)
                self.process_sig.emit('计算进度：{}{:.1f}%'.format('▉▉'*(int((i+1)*100/len(f))//10), (i+1)*100/len(f)))
            self.display_sig.emit('& 存储中......')
            wb.save('./RL-IM.xlsx')
            wb.close()
        except Exception as err:
            self.except_sig2.emit(1)
            self.except_sig.emit('*****************ERROR*****************\n' + str(err) +
                                 '\n*****************************************')
            self.except_sig2.emit(0)
        else:
            self.display_sig.emit('& 处理完成，结果保存在：RL-IM.xlsx')
            self.path_sig.emit('点击查看可直接打开 RL-IM.xlsx')
        finally:
            t2 = self.check_time()
            self.display_sig.emit('& 此次运行共耗时：{:.4f}s'.format(t2 - t1))
            self.flag_sig.emit(True)
            self.quit()


class AlphaCalculate(QThread):
    display_sig = pyqtSignal(str)
    path_sig = pyqtSignal(str)
    flag_sig = pyqtSignal(bool)
    except_sig = pyqtSignal(str)
    except_sig2 = pyqtSignal(int)

    def __init__(self):
        super(AlphaCalculate, self).__init__()

    @staticmethod
    def check_time():
        t = time.perf_counter()
        return t

    def sort_data(self):
        if path.split('.')[-1] == 'xlsx':
            self.display_sig.emit('& 正在读取Excel中的数据......')
            data_wb = load_workbook(path)
            data_ws = data_wb.active
            frequency = [i.value * 1000000000 for i in data_ws['A']]
            real_e = [i.value for i in data_ws['B']]
            imag_e = [i.value for i in data_ws['C']]
            real_u = [i.value for i in data_ws['D']]
            imag_u = [i.value for i in data_ws['E']]
            data_wb.close()
            return frequency, real_e, imag_e, real_u, imag_u
        else:
            with open(path, 'r') as file:
                data = file.readlines()
            self.display_sig.emit('& 正在读取数据......')
            index = 0
            data = [i for i in data if i != '\n']
            while True:
                if len(data[index]) != len(data[index + 3]):
                    index += 1
                elif len(data[index].split()) != len(data[-1].split()):
                    index += 1
                else:
                    break
            frequency = []
            real_e = []
            imag_e = []
            real_u = []
            imag_u = []
            for i in data[index:]:
                if len(data[index].split()[0]) >= 10:
                    frequency.append(float(i.split()[0]))
                else:
                    frequency.append(float(i.split()[0]) * 1000000000)
                real_e.append(float(i.split()[1]))
                imag_e.append(float(i.split()[2]))
                real_u.append(float(i.split()[3]))
                imag_u.append(float(i.split()[4]))
            return frequency, real_e, imag_e, real_u, imag_u

    def run(self):
        self.flag_sig.emit(False)
        t3 = self.check_time()
        try:
            data = self.sort_data()
            f = data[0]
            r_e = data[1]
            i_e = data[2]
            r_u = data[3]
            i_u = data[4]
            self.display_sig.emit('& 初始化结果文件......')
            wb = load_workbook('alpha.xlsx')
            ws = wb['alpha']
            wb.remove(ws)
            new_ws = wb.create_sheet('alpha')
            self.display_sig.emit('& 计算α中，请耐心等待......')
            for i in range(0, len(f)):
                list_alpha = []
                part1 = (math.sqrt(2) * math.pi * f[i]) / 300000000
                part2 = i_u[i] * i_e[i] - r_u[i] * r_e[i]
                part3 = r_u[i] * i_e[i] + i_u[i] * r_e[i]
                alpha = part1 * math.sqrt(
                    part2 + math.sqrt((part2 ** 2) + (part3 ** 2)))
                list_alpha.append(alpha)
                new_ws.append(list_alpha)
            wb.save('./alpha.xlsx')
            wb.close()
        except Exception as err:
            self.except_sig2.emit(1)
            self.except_sig.emit('*****************ERROR*****************\n' + str(err) +
                                 '\n*****************************************')
            self.except_sig2.emit(0)
        else:
            self.display_sig.emit('& 计算结束，结果保存在：alpha.xlsx')
            self.path_sig.emit('点击查看可直接打开 alpha.xlsx')
        finally:
            t4 = self.check_time()
            self.display_sig.emit('& 此次运行共耗时：{:.4f}s'.format(t4 - t3))
            self.flag_sig.emit(True)
            self.quit()


class MyWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.chose_button = QPushButton('选择文件')
        self.c_button1 = QPushButton('计算RL-IM')
        self.c_button2 = QPushButton('计算α')
        self.check_button = QPushButton('查看')
        self.shut_button = QPushButton('关闭')
        self.about_button = QPushButton('关于')
        self.label1 = QLabel('起始厚度(mm)：')
        self.label2 = QLabel('终止厚度(mm)：')
        self.label3 = QLabel('间隔(mm)：')
        self.start_thick = QComboBox()
        self.end_thick = QComboBox()
        self.step = QComboBox()
        self.text = QTextBrowser()
        self.display_text = QTextBrowser()
        self.path_text = QTextBrowser()
        self.a = None
        self.rl_im = None
        self.dic1 = {1: '请选择', 2: '0', 3: '5', 4: '10'}
        self.dic2 = {1: {0: ''},
                     2: {5: '10', 6: '5', 7: '1'},
                     3: {8: '10'},
                     4: {9: '20', 10: '30', 11: '40', 12: '50'}}
        self.dic3 = {1: {0: ''},
                     2: {13: '0.1', 14: '0.05', 15: '0.01'},
                     3: {16: '0.1', 17: '0.05', 18: '0.01'},
                     4: {19: '0.1', 20: '0.2'}}
        self.init_ui()

    def init_ui(self):
        # 主窗口设置
        self.setFont(QFont('Microsoft YaHei', 10))
        self.setFixedSize(550, 630)
        self.setWindowTitle('RL Calculator v2.0.1')
        self.setWindowIcon(QIcon('D:/Pycharm/Project/PyQtProject/pic.ico'))
        self.setStyleSheet('background-color:rgb(240, 248, 255)')
        # 主布局
        layout = QVBoxLayout()
        # 选择文件布局
        layout.addStretch()
        h_layout1 = QHBoxLayout()
        h_layout1.addWidget(self.text)
        h_layout1.addWidget(self.chose_button)
        layout.addLayout(h_layout1)
        # 下拉框布局
        layout.addStretch()
        h_layout8 = QHBoxLayout()
        h_layout6 = QVBoxLayout()
        h_layout7 = QVBoxLayout()
        h_layout6.addWidget(self.label1)
        h_layout6.addWidget(self.label2)
        h_layout6.addWidget(self.label3)
        h_layout7.addWidget(self.start_thick)
        h_layout7.addWidget(self.end_thick)
        h_layout7.addWidget(self.step)
        h_layout8.addStretch()
        h_layout8.addLayout(h_layout6)
        h_layout8.addLayout(h_layout7)
        h_layout8.addStretch()
        layout.addLayout(h_layout8)
        # 计算按钮布局
        layout.addStretch()
        h_layout2 = QHBoxLayout()
        h_layout5 = QVBoxLayout()
        h_layout2.addWidget(self.display_text)
        h_layout5.addWidget(self.c_button1)
        h_layout5.addWidget(self.c_button2)
        h_layout2.addLayout(h_layout5)
        layout.addLayout(h_layout2)
        # 查看文件
        layout.addStretch()
        h_layout3 = QHBoxLayout()
        h_layout3.addWidget(self.path_text)
        h_layout3.addWidget(self.check_button)
        layout.addLayout(h_layout3)
        # 关闭
        layout.addStretch()
        h_layout4 = QHBoxLayout()
        h_layout4.addStretch()
        h_layout4.addWidget(self.about_button)
        h_layout4.addWidget(self.shut_button)
        h_layout4.addStretch()
        layout.addLayout(h_layout4)
        self.setLayout(layout)
        # 选择类样式
        self.chose_button.setFixedSize(100, 80)
        self.chose_button.setStyleSheet('background-color:rgb(230, 230, 250)')
        self.text.setFixedSize(360, 105)
        self.text.setStyleSheet('background-color:rgb(253, 245, 230)')
        self.text.setPlaceholderText('可直接打开的文件后缀：\n'
                                     '#  (*.dat)、(*.prn)、(*.txt)\n'
                                     '自定义的数据文件后缀：\n'
                                     '#  (*.xlsx)')
        # 计算类样式
        self.c_button1.setFixedSize(100, 80)
        self.c_button2.setFixedSize(100, 80)
        self.c_button1.setStyleSheet('background-color:rgb(151, 255, 255)')
        self.c_button2.setStyleSheet('background-color:rgb(151, 255, 255)')
        # 展示类样式
        self.display_text.setFixedSize(360, 270)
        self.display_text.setTextColor(QColor(0, 0, 139))
        self.display_text.setStyleSheet('background-color:rgb(253, 245, 230)')
        self.display_text.setPlaceholderText('注意事项：\n'
                                             '# 计算开始前请先关闭 RL-IM.xlsx ！！！\n'
                                             '# 若出错或闪退，请将数据复制到Excel表中\n'
                                             '   -以第一个单元格为开始(f,ε′,ε″,μ′,μ″)\n'
                                             '   -只复制数据，不要复制标签等无关内容！\n'
                                             '   -频率单位为（GHz）\n'
                                             '# 计算结束请及时另存数据！！！')
        # 查看类样式
        self.check_button.setFixedSize(100, 35)
        self.check_button.setStyleSheet('background-color:rgb(230, 230, 250)')
        self.path_text.setFixedSize(360, 35)
        self.path_text.setTextColor(QColor(0, 0, 139))
        self.path_text.setStyleSheet('background-color:rgb(253, 245, 230)')
        self.path_text.setPlaceholderText('数据文件')
        # 下拉框样式
        self.start_thick.setFixedSize(100, 30)
        self.end_thick.setFixedSize(100, 30)
        self.step.setFixedSize(100, 30)
        self.label1.setAlignment(Qt.AlignRight)
        self.label2.setAlignment(Qt.AlignRight)
        self.label3.setAlignment(Qt.AlignRight)
        for (keys, value) in self.dic1.items():
            self.start_thick.addItem(value, QVariant(keys))
        # 按钮点击事件
        self.chose_button.clicked.connect(self.findfile)

        self.c_button1.clicked.connect(self.click_rl)
        self.c_button2.clicked.connect(self.click_alpha)

        self.check_button.clicked.connect(self.openfile)

        self.about_button.clicked.connect(self.click_about)
        self.shut_button.clicked.connect(QApplication.quit)

        self.start_thick.activated.connect(self.on_end_thick)

    def on_end_thick(self, key):
        self.end_thick.clear()
        self.step.clear()
        data = self.start_thick.itemData(key)
        for (keys, value) in self.dic2[data].items():
            self.end_thick.addItem(value, QVariant(keys))
        for (keys, value) in self.dic3[data].items():
            self.step.addItem(value, QVariant(keys))

    def findfile(self):
        findfile_name = QFileDialog.getOpenFileName(self, '选择文件',
                                                    '',
                                                    'Data files(*.dat *.prn *.xlsx *.txt)')
        tex = str(findfile_name[0])
        self.text.setText(tex)
        self.path_text.clear()
        self.display_text.clear()
        global path
        path = tex

    def check_file1(self, filename):
        self.display_text.clear()
        self.path_text.clear()
        self.display_text.setText('& 检查文件......')
        if not os.path.exists(filename):
            wb = Workbook()
            wb.remove(wb['Sheet'])
            wb.create_sheet('RL')
            wb.create_sheet('IM')
            wb.save(filename)
            wb.close()
        else:
            pass

    def check_file2(self, filename):
        self.display_text.clear()
        self.path_text.clear()
        self.display_text.setText('& 检查文件......')
        if not os.path.exists(filename):
            wb = Workbook()
            wb.remove(wb['Sheet'])
            wb.create_sheet('alpha')
            wb.save(filename)
            wb.close()
        else:
            pass

    def openfile(self):
        str_path = self.path_text.toPlainText()
        os.startfile(str_path.split(' ')[-1])

    def click_about(self):
        msg = 'Version：2.0.1<br>' \
              'Author：Xuetao Sun<br>Download：' \
              '<a href="https://github.com/MrSunxt/Reflection-Loss-Calculator/releases">GitHub</a><br>' \
              'Tutorial：<a href="https://www.bilibili.com/video/BV15A411S7BP/?vd_source=b29549a87424fe11cefd502a92abc9d7">' \
              'B站</a><br>' \
              'Feedback：18268219283@163.com'
        QMessageBox.about(self, 'About', msg)

    def click_rl(self):
        if path == '':
            self.display_text.setText('\n\n\n\n{}'
                                      .format('                  ！！！ 空文件 ！！！\n'
                                              * 3))
        elif self.step.currentText() == '':
            self.display_text.setText('\n\n\n\n{}'
                                      .format('                ！！！ 请选择厚度 ！！！\n'
                                              * 3))
        else:
            global start_thick, end_thick, thick, step, thick_count
            start_thick = int(self.start_thick.currentText())
            end_thick = int(self.end_thick.currentText())
            thick = end_thick - start_thick
            step = float(self.step.currentText())
            thick_count = int(thick / step + 1)
            self.check_file1('RL-IM.xlsx')
            self.rl_im = RlCalculate()
            self.rl_im.flag_sig.connect(self.enable_button)
            self.rl_im.flag_sig.connect(self.button_type1)
            self.rl_im.display_sig.connect(self.displaytext)
            self.rl_im.process_sig.connect(self.p_text)
            self.rl_im.except_sig2.connect(self.display_color)
            self.rl_im.except_sig.connect(self.except_msg)
            self.rl_im.path_sig.connect(self.p_text)
            self.rl_im.start()

    def click_alpha(self):
        if path == '':
            self.display_text.setText('\n\n\n\n{}'
                                      .format('                  ！！！ 空文件 ！！！\n'
                                              * 3))
        else:
            self.check_file2('alpha.xlsx')
            self.a = AlphaCalculate()
            self.a.flag_sig.connect(self.enable_button)
            self.a.flag_sig.connect(self.button_type2)
            self.a.display_sig.connect(self.displaytext)
            self.a.except_sig2.connect(self.display_color)
            self.a.except_sig.connect(self.except_msg)
            self.a.path_sig.connect(self.p_text)
            self.a.start()

    def except_msg(self, message):
        self.display_text.append('\n' + message + '\n')

    def display_color(self, flag):
        if flag == 1:
            self.display_text.setTextColor(QColor(165, 42, 42))
        else:
            self.display_text.setTextColor(QColor(0, 0, 139))

    def displaytext(self, t):
        self.display_text.append(t)

    def p_text(self, p):
        self.path_text.setText(p)

    def enable_button(self, flag):
        self.chose_button.setEnabled(flag)
        self.c_button1.setEnabled(flag)
        self.c_button2.setEnabled(flag)
        self.shut_button.setEnabled(flag)
        self.start_thick.setEnabled(flag)
        self.end_thick.setEnabled(flag)
        self.step.setEnabled(flag)

    def button_type1(self, flag):
        if not flag:
            self.c_button1.setStyleSheet('background-color:rgb(253, 245, 230)')
            self.c_button1.setFont(QFont('Microsoft YaHei', 10))
        else:
            self.c_button1.setStyleSheet('background-color:rgb(151, 255, 255)')

    def button_type2(self, flag):
        if not flag:
            self.c_button2.setStyleSheet('background-color:rgb(253, 245, 230)')
            self.c_button2.setFont(QFont('Microsoft YaHei', 10))
        else:
            self.c_button2.setStyleSheet('background-color:rgb(151, 255, 255)')


if __name__ == '__main__':
    app = QApplication(sys.argv)

    window = MyWindow()
    window.show()

    app.exec()
