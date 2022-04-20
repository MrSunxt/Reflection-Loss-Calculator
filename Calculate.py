import sys
import os
import time
import cmath

from PyQt5.QtGui import QIcon, QColor
from openpyxl import load_workbook, Workbook
from PyQt5.QtWidgets import *


class MyWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()

    def init_ui(self):
        self.resize(500, 400)
        self.setWindowTitle('反射损耗计算器v1.0-bysxt')
        self.setWindowIcon(QIcon('./pic.png'))
        self.setStyleSheet('background-color:rgb(240, 248, 255)')
        layout = QVBoxLayout()
        layout.addStretch()

        h_layout1 = QHBoxLayout()
        chose_button = QPushButton('选择文件')
        chose_button.setFixedSize(100, 75)
        chose_button.setStyleSheet('background-color:rgb(230, 230, 250)')
        chose_button.clicked.connect(self.findfile)
        self.text = QTextEdit()
        self.text.setPlaceholderText('文件路径')
        self.text.setFixedSize(400, 75)
        self.text.setStyleSheet('background-color:rgb(253, 245, 230)')
        h_layout1.addWidget(self.text)
        h_layout1.addWidget(chose_button)
        layout.addLayout(h_layout1)

        layout.addStretch()
        h_layout2 = QHBoxLayout()
        c_button1 = QPushButton('计算RL')
        c_button2 = QPushButton('计算IM')
        c_button1.setFixedSize(100, 50)
        c_button2.setFixedSize(100, 50)
        c_button1.setStyleSheet('background-color:rgb(127, 255, 212)')
        c_button2.setStyleSheet('background-color:rgb(127, 255, 212)')
        c_button1.clicked.connect(self.calculate_rl)
        c_button2.clicked.connect(self.calculate_im)
        h_layout2.addStretch()
        h_layout2.addWidget(c_button1)
        h_layout2.addStretch()
        h_layout2.addWidget(c_button2)
        h_layout2.addStretch()
        layout.addLayout(h_layout2)

        layout.addStretch()
        self.display_text = QTextEdit()
        hold_str = '1.选择一个.dat文件\n2.点击计算RL或IM(默认计算0~10mm)\n' \
                   '3.查看计算好的excel文件并自行另存'
        self.display_text.setPlaceholderText(hold_str)
        self.display_text.setFixedSize(500, 75)
        self.display_text.setTextColor(QColor(255, 0, 0))
        self.display_text.setStyleSheet('background-color:rgb(253, 245, 230)')
        layout.addWidget(self.display_text)

        layout.addStretch()
        h_layout3 = QHBoxLayout()
        check_button = QPushButton('查看')
        self.path_line = QLineEdit()
        self.path_line.setPlaceholderText('数据文件路径')
        check_button.setFixedSize(100, 50)
        self.path_line.setFixedSize(400, 50)
        check_button.setStyleSheet('background-color:rgb(230, 230, 250)')
        self.path_line.setStyleSheet('background-color:rgb(253, 245, 230)')
        check_button.clicked.connect(self.openfile)
        h_layout3.addWidget(self.path_line)
        h_layout3.addWidget(check_button)
        layout.addLayout(h_layout3)

        layout.addStretch()
        h_layout4 = QHBoxLayout()
        shut_button = QPushButton('关闭')
        shut_button.setStyleSheet('background-color:rgb(255, 246, 143)')
        shut_button.clicked.connect(QApplication.quit)
        h_layout4.addStretch()
        h_layout4.addWidget(shut_button)
        h_layout4.addStretch()
        layout.addLayout(h_layout4)

        self .setLayout(layout)
        self.show()

    def findfile(self):
        findfile_name = QFileDialog.getOpenFileName(self, '选择文件',
                                                    '',
                                                    'Data files(*.dat)')
        tex = str(findfile_name[0])
        self.text.setText(tex)

    def calculate_rl(self):
        t1 = time.time()
        path = self.text.toPlainText()
        f = open(str(path), 'r')
        txt = f.read()
        f.close()
        start = txt.find(' 1.000000')
        data = txt[start:].split()
        data = [float(i) for i in data]

        def sort(i):
            name = []
            count = i
            while count < 8005:
                j = data[count]
                name.append(j)
                count += 5
            return name

        frequency = sort(0)
        real_e = sort(1)
        imag_e = sort(2)
        real_u = sort(3)
        imag_u = sort(4)
        e = []
        u = []
        for i in range(0, 1601):
            e.append(complex(real_e[i], -imag_e[i]))
            u.append(complex(real_u[i], -imag_u[i]))

        wb = load_workbook('RL.xlsx')
        ws = wb['Sheet1']
        wb.remove(ws)
        new_ws = wb.create_sheet('Sheet1')

        for i in range(0, 1601):
            list_rl = []
            for j in range(0, 101):
                D = 0.0001 * j
                d = round(D, 4)
                Zin = cmath.sqrt(u[i] / e[i]) * cmath.tanh(complex(0, 1) * ((2 * cmath.pi * frequency[i] * 1000000000 * d) / 300000000) * cmath.sqrt(u[i] * e[i]))
                Rl = 20 * cmath.log10(abs((Zin - 1) / (Zin + 1)))
                list_rl.append(Rl.real)
            new_ws.append(list_rl)
        wb.save('./RL.xlsx')
        wb.close()
        t2 = time.time()
        f_path = 'RL.xlsx'
        self.path_line.setText(f_path)
        self.display_text.setText('运行结束，耗时：{}s\n请及时保存数据\n'
                                  '(点击查看并另存文件)'
                                  .format(t2 - t1))

    def calculate_im(self):
        t3 = time.time()
        path = self.text.toPlainText()
        f = open(str(path), 'r')
        txt = f.read()
        f.close()
        start = txt.find(' 1.000000')
        data = txt[start:].split()
        data = [float(i) for i in data]

        def sort(i):
            name = []
            count = i
            while count < 8005:
                j = data[count]
                name.append(j)
                count += 5
            return name

        frequency = sort(0)
        real_e = sort(1)
        imag_e = sort(2)
        real_u = sort(3)
        imag_u = sort(4)
        e = []
        u = []
        for i in range(0, 1601):
            e.append(complex(real_e[i], -imag_e[i]))
            u.append(complex(real_u[i], -imag_u[i]))

        wb = load_workbook('IM.xlsx')
        ws = wb['Sheet1']
        wb.remove(ws)
        new_ws = wb.create_sheet('Sheet1')

        for i in range(0, 1601):
            list_im = []
            for j in range(0, 101):
                D = 0.0001 * j
                d = round(D, 4)
                zin = cmath.sqrt(u[i] / e[i]) * cmath.tanh(complex(0, 1) * ((2 * cmath.pi * frequency[i] * 1000000000 * d) / 300000000) * cmath.sqrt(u[i] * e[i]))
                im = abs(zin)
                list_im.append(im)
            new_ws.append(list_im)
        wb.save('./IM.xlsx')
        wb.close()
        t4 = time.time()
        f_path = 'IM.xlsx'
        self.path_line.setText(f_path)
        self.display_text.setText('运行结束，耗时：{}s\n'
                                  '请及时保存数据\n'
                                  '(点击查看并另存文件)'
                                  .format(t4 - t3))

    def openfile(self):
        os.startfile(self.path_line.text())


if __name__ == '__main__':
    app = QApplication(sys.argv)

    window = MyWindow()

    app.exec()
